# -*- coding: utf-8 -*-
from copy import copy
import datetime
import io
import os
import re
import zipfile

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
import pandas as pd
from pythainlp.util import bahttext
import streamlit as st

from doc_processor import (
    format_budget_money,
    format_thai_date,
    process_docx,
    to_thai_num,
)
from offer_components import render_person_inputs
from offer_helpers import (
    add_business_days,
    load_shops_data,
    load_teacher_data,
)
from offer_modals import add_shop_modal


def generate_hiring_space_excel_internal(
    valid_items,
    total_amount,
    project_name="",
    department="",
    budget_type="",
    parcel_no="",
    receiver="",
    receiver_sub="",
):
    """ฟังก์ชันสร้างไฟล์ Excel ข้อกำหนดการจ้าง (Space) แบบอัตโนมัติภายในตัว"""
    template_path = (
        "space.xlsx"
        if os.path.exists("space.xlsx")
        else os.path.join("templates_4color", "space.xlsx")
    )

    if os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "รายการพัสดุและสเปค"
        ws.append(["ลำดับ", "รายการพัสดุ / รายละเอียดสเปค", "จำนวน", "หน่วย"])
        for idx, item in valid_items.reset_index(drop=True).iterrows():
            ws.append(
                [
                    idx + 1,
                    item.get("รายการพัสดุ / รายละเอียดสเปค", ""),
                    item.get("จำนวน", 1),
                    item.get("หน่วย", "งาน"),
                ]
            )
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    try:
        budget_text = bahttext(total_amount)
    except Exception:
        budget_text = ""

    budget_with_text = (
        f"{total_amount:,.2f} บาท ({budget_text})" if total_amount > 0 else ""
    )

    mapping = {
        "{{BUDGET_WITH_TEXT}}": budget_with_text,
        "{{budget_with_text}}": budget_with_text,
        "{{BUDGET_TEXT_MID}}": budget_text,
        "{{budget_text_mid}}": budget_text,
        "{{receiver}}": str(receiver),
        "{{receiver_sub}}": str(receiver_sub),
        "{{total_amount}}": f"{total_amount:,.2f}",
        "{{DEPARTMENT}}": str(department),
        "{{department}}": str(department),
        "{{BUDGET_TYPE}}": str(budget_type),
        "{{budget_type}}": str(budget_type),
        "{{PARCEL_NO}}": str(parcel_no),
        "{{parcel_no}}": str(parcel_no),
    }

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    empty_border = Border()

    for ws in wb.worksheets:
        start_row = None
        for r in range(1, ws.max_row + 1):
            val_col1 = str(ws.cell(row=r, column=1).value or "")
            val_col2 = str(ws.cell(row=r, column=2).value or "")
            if "{% for item in items %}" in val_col1 or "{{item.name}}" in val_col2:
                start_row = r
                break

        if start_row:
            num_items = len(valid_items)
            max_cols = 4

            template_styles = {}
            for c in range(1, max_cols + 1):
                src_cell = ws.cell(row=start_row, column=c)
                template_styles[c] = {
                    "font": copy(src_cell.font),
                    "alignment": copy(src_cell.alignment),
                }

            merged_to_shift = []
            merged_to_remove = []
            for rng in list(ws.merged_cells.ranges):
                if rng.min_row > start_row:
                    merged_to_shift.append(
                        (rng.min_row, rng.min_col, rng.max_row, rng.max_col)
                    )
                    merged_to_remove.append(str(rng))

            for rng_str in merged_to_remove:
                ws.unmerge_cells(rng_str)

            num_page_breaks = (num_items - 1) // 19 if num_items > 19 else 0
            shift_amount = (num_items - 1) + (num_page_breaks * 2)

            if shift_amount > 0:
                ws.insert_rows(start_row + 1, amount=shift_amount)

            for min_r, min_c, max_r, max_c in merged_to_shift:
                ws.merge_cells(
                    start_row=min_r + shift_amount,
                    start_column=min_c,
                    end_row=max_r + shift_amount,
                    end_column=max_c,
                )

            current_row = start_row
            for idx, item in valid_items.reset_index(drop=True).iterrows():
                item_name = str(item.get("รายการพัสดุ / รายละเอียดสเปค", ""))
                item_qty = item.get("จำนวน", 1)
                item_unit = str(item.get("หน่วย", "งาน"))

                ws.cell(row=current_row, column=1, value=idx + 1)
                ws.cell(row=current_row, column=2, value=item_name)
                ws.cell(row=current_row, column=3, value=item_qty)
                ws.cell(row=current_row, column=4, value=item_unit)

                for col in range(1, 5):
                    cell = ws.cell(row=current_row, column=col)
                    cell.border = thin_border
                    if template_styles.get(col):
                        cell.font = copy(template_styles[col]["font"])
                        cell.alignment = copy(template_styles[col]["alignment"])

                ws.cell(row=current_row, column=1).alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                ws.cell(row=current_row, column=3).alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                ws.cell(row=current_row, column=4).alignment = Alignment(
                    horizontal="center", vertical="center"
                )

                if (idx + 1) % 19 == 0 and (idx + 1) < num_items:
                    current_row += 1
                    for col in range(1, 5):
                        ws.cell(row=current_row, column=col).border = empty_border

                    current_row += 1
                    next_page_num = ((idx + 1) // 19) + 1

                    ws.merge_cells(
                        start_row=current_row,
                        start_column=1,
                        end_row=current_row,
                        end_column=max_cols,
                    )
                    cell_continue = ws.cell(row=current_row, column=1)
                    cell_continue.value = f"(ต่อแผ่น {next_page_num})"
                    cell_continue.alignment = Alignment(
                        horizontal="right", vertical="center"
                    )

                    for col in range(1, max_cols + 1):
                        ws.cell(row=current_row, column=col).border = empty_border

                    if template_styles.get(1):
                        base_font = copy(template_styles[1]["font"])
                        base_font.italic = True
                        base_font.bold = True
                        cell_continue.font = base_font

                    ws.row_breaks.append(Break(id=current_row))

                current_row += 1

        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell_obj = ws.cell(row=r, column=c)
                if type(cell_obj).__name__ != "MergedCell":
                    val = str(cell_obj.value or "")
                    if "{{PROJECT_NAME}}" in val or "{{project_name}}" in val:
                        cell_obj.value = val.replace(
                            "{{PROJECT_NAME}}", str(project_name)
                        ).replace("{{project_name}}", str(project_name))
                        current_align = cell_obj.alignment
                        cell_obj.alignment = Alignment(
                            horizontal=current_align.horizontal or "left",
                            vertical=current_align.vertical or "center",
                            wrap_text=True,
                        )
                    else:
                        for key, target in mapping.items():
                            if key in val:
                                cell_obj.value = val.replace(key, target)

        max_col_letter = get_column_letter(ws.max_column)
        ws.print_area = f"A1:{max_col_letter}{ws.max_row}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _extract_shop_info(row):
    """Helper ดึงข้อมูล ที่อยู่, เบอร์โทร, เลขภาษี จาก row ของ df_shops"""

    def find_val(candidates):
        for cand in candidates:
            for c in row.index:
                if str(c).lower().strip() == cand.lower():
                    val = row[c]
                    if pd.notna(val) and str(val).strip().lower() not in [
                        "nan",
                        "none",
                        "",
                    ]:
                        return str(val).strip()
        return ""

    return (
        find_val(["address", "shop_address", "ที่อยู่", "addr"]),
        find_val(["phone", "tel", "เบอร์โทร", "เบอร์โทรศัพท์", "telephone"]),
        find_val(["tax_id", "taxid", "เลขประจำตัวผู้เสียภาษี", "tax_no", "tax"]),
    )


def reset_hiring_form():
    """ฟังก์ชันสำหรับล้างข้อมูลในฟอร์มจัดจ้าง"""
    keys_to_clear = [
        "hiring_parcel_no",
        "hiring_project_name",
        "hiring_budget",
        "hiring_item_count",
        "hiring_dept",
        "hiring_doc_no",
        "hiring_doc_save",
        "hiring_budget_mid",
        "hiring_sub_grant",
        "hiring_sub_budget",
        "hiring_submit_no",
        "hiring_items_editor",
    ]
    for key in keys_to_clear:
        if key in st.session_state:
            del st.session_state[key]

    st.toast("ล้างข้อมูลเรียบร้อยแล้ว!", icon="🧹")


def on_edit_vendor_change_hiring(df_shops):
    selected_vendor = st.session_state.get("dialog_edit_hiring_vendor_select")
    if selected_vendor and not df_shops.empty:
        match = df_shops[df_shops["clean_shop_name"] == str(selected_vendor).strip()]
        if not match.empty:
            addr, phone, tax = _extract_shop_info(match.iloc[0])
            st.session_state["dialog_edit_hiring_vendor_addr"] = addr
            st.session_state["dialog_edit_hiring_vendor_phone"] = phone
            st.session_state["dialog_edit_hiring_vendor_tax"] = tax


@st.dialog("✏️ แก้ไขข้อมูลผู้รับจ้าง / ร้านค้า")
def edit_hiring_shop_dialog(shop_list_options, df_shops):
    if not shop_list_options:
        st.warning("ยังไม่มีข้อมูลผู้รับจ้างในระบบ")
        return

    if "dialog_edit_hiring_vendor_select" not in st.session_state:
        st.session_state["dialog_edit_hiring_vendor_select"] = shop_list_options[0]
        on_edit_vendor_change_hiring(df_shops)

    st.selectbox(
        "เลือกชื่อผู้รับจ้าง / บริษัท ที่ต้องการแก้ไข",
        options=shop_list_options,
        key="dialog_edit_hiring_vendor_select",
        on_change=on_edit_vendor_change_hiring,
        args=(df_shops,),
    )
    st.text_area("ที่อยู่", key="dialog_edit_hiring_vendor_addr")
    st.text_input("เบอร์โทรศัพท์", key="dialog_edit_hiring_vendor_phone")
    st.text_input("เลขประจำตัวผู้เสียภาษี", key="dialog_edit_hiring_vendor_tax")

    if st.button("💾 บันทึกการแก้ไข", type="primary", use_container_width=True):
        selected_vendor = st.session_state.get("dialog_edit_hiring_vendor_select")
        st.success(f"บันทึกการแก้ไขผู้รับจ้าง '{selected_vendor}' เรียบร้อยแล้ว")
        st.rerun()


def render_hiring_page():
    st.markdown(
        """
        <style>
            @import url('https://fonts.googleapis.com/css2?family=TH+Sarabun+New:wght@400;600;700&display=swap');
            @import url('https://fonts.cdnfonts.com/css/th-sarabunpsk');

            header[data-testid="stHeader"] { display: none !important; }
            .main .block-container { padding-top: 1rem !important; padding-bottom: 2.5rem !important; max-width: 1100px !important; }
            html, body, [class*="css"], .stMarkdown, .stText, p, label, input, select, textarea, button, span, div { font-family: 'TH SarabunPSK', 'TH Sarabun New', 'Sarabun', sans-serif !important; font-size: 103.5% !important; }
            div[data-testid="stHorizontalBlock"] { align-items: flex-start !important; gap: 1rem !important; }
            .stTextInput label, .stSelectbox label, .stDateInput label, .stRadio label { font-weight: 600 !important; font-size: 105% !important; margin-bottom: 4px !important; color: #334155 !important; }
            div[data-baseweb="select"] *, div[data-baseweb="input"] input { font-family: 'TH SarabunPSK', 'TH Sarabun New', 'Sarabun', sans-serif !important; font-size: 105% !important; }
            div[data-testid="stDateInput"] input { font-family: 'TH SarabunPSK', 'TH Sarabun New', 'Sarabun', sans-serif !important; font-size: 105% !important; line-height: 44px !important; height: 44px !important; padding-top: 0px !important; padding-bottom: 0px !important; }
            div[data-baseweb="input"] > div, div[data-baseweb="select"] > div { min-height: 44px !important; height: 44px !important; align-items: center !important; border-radius: 6px !important; background-color: #F8FAFC !important; border: 1px solid #CBD5E1 !important; }
            
            div[data-baseweb="textarea"] > div { min-height: 44px !important; border-radius: 6px !important; background-color: #F8FAFC !important; border: 1px solid #CBD5E1 !important; }
            div[data-baseweb="textarea"] textarea { font-family: 'TH SarabunPSK', 'TH Sarabun New', 'Sarabun', sans-serif !important; font-size: 95% !important; line-height: 1.2 !important; padding: 6px 8px !important; }

            h1 { font-size: 3.5rem !important; font-weight: 700 !important; margin-top: 0rem !important; margin-bottom: 0.5rem !important; color: #0F172A !important; }
            h3, .stSubheader { font-size: 2.6rem !important; font-weight: 700 !important; color: #0F172A !important; margin-top: 0.8rem !important; margin-bottom: 0.6rem !important; border-bottom: 2px solid #CBD5E1; padding-bottom: 4px; }
            h5 { font-size: 1.8rem !important; font-weight: 600 !important; color: #1E293B !important; margin-top: 0.2rem !important; margin-bottom: 0.5rem !important; background-color: #F1F5F9; padding: 4px 10px; border-left: 4px solid #2563EB; border-radius: 0 4px 4px 0; }
            .stButton > button { padding: 0.5rem 1.5rem !important; font-size: 100% !important; border-radius: 8px !important; font-weight: 600 !important; }
            .stElementContainer { margin-bottom: 0.4rem !important; }
            
            .header-btn-container { margin-top: 25px; }
            .header-btn-container button { height: 44px !important; min-height: 44px !important; width: 100% !important; }
        </style>
        """,
        unsafe_allow_html=True,
    )

    st.title("🔨 ระบบสร้างเอกสารจัดจ้าง")

    # --- กำหนดค่าเริ่มต้น State สำหรับตารางรายการพัสดุ (ป้องกันข้อมูลหลุด/ต้องกรอกซ้ำ) ---
    if "hiring_items_editor" not in st.session_state:
        st.session_state["hiring_items_editor"] = pd.DataFrame(
            [
                {
                    "รายการพัสดุ / รายละเอียดสเปค": "",
                    "จำนวน": 1,
                    "หน่วย": "งาน",
                }
            ]
        )

    # --- แถวบนสุด ---
    col_top1, col_top2 = st.columns([1, 2], gap="medium")
    with col_top1:
        if st.button("⬅️ กลับหน้าหลัก", use_container_width=True):
            st.session_state.page = "home"
            st.rerun()

    with col_top2:
        if "hiring_parcel_no" not in st.session_state:
            st.session_state["hiring_parcel_no"] = ""
        else:
            val_init = st.session_state["hiring_parcel_no"]
            if "/" in val_init or "_" in val_init:
                st.session_state["hiring_parcel_no"] = val_init.replace(
                    "/", "-"
                ).replace("_", "-")

        def update_hiring_parcel():
            val = st.session_state["hiring_parcel_no"]
            if "/" in val or "_" in val:
                st.session_state["hiring_parcel_no"] = val.replace("/", "-").replace(
                    "_", "-"
                )

        parcel_no = st.text_input(
            "",
            placeholder="ตัวอย่าง: เลขพัสดุ (เช่น 317-69 หรือ 111-69)",
            key="hiring_parcel_no",
            on_change=update_hiring_parcel,
        )

    TEMPLATE_DIR = "templates_hiring"
    os.makedirs(TEMPLATE_DIR, exist_ok=True)
    template_files = sorted(
        [
            f
            for f in os.listdir(TEMPLATE_DIR)
            if f.endswith(".docx") and not f.startswith("~$")
        ]
    )

    # --- ส่วนที่ 1: ข้อมูลการจัดจ้างหลัก ---
    st.subheader("📝 1. ข้อมูลการจัดจ้าง (หน้า ส.1)")
    col1, col2 = st.columns([1, 1], gap="large")

    with col1:
        project_name = st.text_input(
            "ชื่อโครงการ / ชื่องานจ้าง",
            placeholder="ตัวอย่าง: จ้างซ่อมแซมเครื่องปรับอากาศ",
            key="hiring_project_name",
        )
        budget = st.text_input(
            "จำนวนเงิน / วงเงินที่จะจ้าง (พิมพ์เฉพาะตัวเลข)",
            placeholder="ตัวอย่าง: 15000",
            key="hiring_budget",
        )

        clean_num = 0.0
        try:
            clean_num = float(
                budget.replace(",", "").replace(" ", "").replace("บาท", "")
            )
            budget_text = bahttext(clean_num)
        except ValueError:
            budget_text = budget

        st.info(f"💡 **แปลงเป็นตัวหนังสืออัตโนมัติ:** {budget_text}")

        item_count_raw = st.text_input(
            "จำนวนรายการ / งานทั้งหมด (เช่น 1 งาน)",
            placeholder="ตัวอย่าง: 1 งาน",
            key="hiring_item_count",
        )

        clean_max_item_count = 0
        try:
            match_cnt = re.search(r"\d+", item_count_raw.replace(",", ""))
            if match_cnt:
                clean_max_item_count = int(match_cnt.group())
        except ValueError:
            clean_max_item_count = 0

        budget_source_type = st.radio(
            "แหล่งเงินงบประมาณ",
            ["บกศ. (เงินรายได้สถานศึกษา)", "เงินอุดหนุน", "งปม. (เงินงบประมาณ)"],
            key="hiring_budget_source",
        )

        budget_type_text = ""
        if budget_source_type == "บกศ. (เงินรายได้สถานศึกษา)":
            budget_type_text = "เงินรายได้ของสถานศึกษา (บกศ.)"
        elif budget_source_type == "เงินอุดหนุน":
            sub_detail = st.text_input(
                "ระบุรายละเอียดเงินอุดหนุน",
                value="กิจกรรมพัฒนาผู้เรียน",
                placeholder="ตัวอย่าง: ค่ากิจกรรมพัฒนาผู้เรียน / ค่าเรียนฟรี 15 ปี",
                key="hiring_sub_grant",
            )
            budget_type_text = f"เงินอุดหนุน {sub_detail}"
        elif budget_source_type == "งปม. (เงินงบประมาณ)":
            sub_detail = st.text_input(
                "ระบุรายละเอียดเงินงบประมาณ",
                value="เงินอุดหนุน",
                placeholder="ตัวอย่าง: งบอุดหนุน / งบดำเนินงาน",
                key="hiring_sub_budget",
            )
            budget_type_text = f"เงินงบประมาณ {sub_detail}"

    with col2:
        department = st.text_input(
            "งาน หรือ แผนกวิชาที่จัดทำ",
            placeholder="ตัวอย่าง: แผนกวิชาช่างยนต์ / งานกิจกรรมนักเรียนนักศึกษา",
            key="hiring_dept",
        )
        selected_date = st.date_input(
            "วันที่เอกสาร (หน้า ส.1)", datetime.date.today(), key="hiring_doc_date"
        )
        doc_no_raw = st.text_input(
            "เลขที่คำสั่ง", placeholder="ตัวอย่าง: 961/2569", key="hiring_doc_no"
        )
        doc_no_save = st.text_input(
            "เลขที่บันทึก", placeholder="ตัวอย่าง: 961", key="hiring_doc_save"
        )

    use_thai_num, use_thai_num2 = True, True

    # --- ส่วนที่ 2: คณะกรรมการ ---
    st.write("")
    st.subheader("👥 2. คำสั่งคณะกรรมการจัดจ้าง / ตรวจรับ")
    person_options, person_dict = load_teacher_data("teachers.xlsx")

    col_count1, col_count2 = st.columns([1, 1], gap="large")
    buy_count = col_count1.selectbox(
        "จำนวนกรรมการจัดจ้าง",
        options=[1, 2, 3],
        index=2,
        key="hiring_buy_count_select",
    )
    check_count = col_count2.selectbox(
        "จำนวนกรรมการตรวจรับ",
        options=[1, 2, 3],
        index=2,
        key="hiring_check_count_select",
    )

    # 1. คณะกรรมการจัดจ้าง
    st.markdown("##### 🛠️ คณะกรรมการจัดจ้าง")
    buy_persons = []
    defaults_buy = [
        ("จัดจ้าง 1", "นายสมชาย ใจดี", "ประธานกรรมการฯ"),
        ("จัดจ้าง 2", "นางสาวสมหญิง ใจงาม", "กรรมการฯ"),
        ("จัดจ้าง 3", "นายสมปอง ใจกล้า", "กรรมการฯ"),
    ]
    with st.container(border=True):
        for idx in range(3):
            label, def_name, def_pos = defaults_buy[idx]
            if idx < buy_count:
                p = render_person_inputs(
                    label,
                    f"hiring_buy{idx+1}",
                    def_name,
                    "",
                    def_pos,
                    person_options,
                    person_dict,
                )
            else:
                p = ("", "", "")
            buy_persons.append(p)

    # 2. คณะกรรมการตรวจรับ
    st.markdown("##### 🔍 คณะกรรมการตรวจรับงานจ้าง")
    check_persons = []
    check_pos_options = (
        ["ผู้ตรวจรับการจ้าง"]
        if check_count == 1
        else ["ประธานกรรมการฯ", "กรรมการฯ", "กรรมการและเลขานุการฯ"]
    )
    default_check_pos1 = "ผู้ตรวจรับการจ้าง" if check_count == 1 else "ประธานกรรมการฯ"
    defaults_check = [
        ("ตรวจรับ 1", "นายสมชาย ใจดี", default_check_pos1),
        ("ตรวจรับ 2", "นางสาวสมหญิง ใจงาม", "กรรมการฯ"),
        ("ตรวจรับ 3", "นายสมปอง ใจกล้า", "กรรมการฯ"),
    ]

    with st.container(border=True):
        for idx in range(3):
            label, def_name, def_pos = defaults_check[idx]
            if idx < check_count:
                kw = {"pos_options_custom": check_pos_options} if idx == 0 else {}
                p = render_person_inputs(
                    label,
                    f"hiring_check{idx+1}",
                    def_name,
                    "",
                    def_pos,
                    person_options,
                    person_dict,
                    **kw,
                )
            else:
                p = ("", "", "")
            check_persons.append(p)

    # --- ส่วนที่ 3: บันทึกรายงานผลการพิจารณา ---
    st.write("")
    st.subheader("📋 3. บันทึกรายงานผลการพิจารณา")

    df_shops = load_shops_data()
    shop_list_options = []
    if not df_shops.empty:
        name_col = next(
            (
                col
                for col in df_shops.columns
                if str(col).lower()
                in [
                    "shop_name",
                    "name",
                    "shopname",
                    "ร้านค้า",
                    "ชื่อร้าน",
                    "ผู้รับจ้าง",
                ]
            ),
            df_shops.columns[0],
        )
        df_shops["clean_shop_name"] = df_shops[name_col].astype(str).str.strip()
        shop_list_options = [
            s
            for s in df_shops["clean_shop_name"].unique()
            if s and s.lower() not in ["nan", "none", "nat", ""]
        ]

    top_col1, top_col2, top_col3, top_col4 = st.columns([1.2, 1, 1, 1], gap="small")
    selected_date2 = top_col1.date_input(
        "วันที่รายงานผล", datetime.date.today(), key="hiring_date_report"
    )
    shop_count = top_col2.selectbox(
        "จำนวนผู้รับจ้าง/บริษัท",
        options=[1, 2, 3, 4],
        index=0,
        key="hiring_shop_count_select",
    )

    with top_col3:
        st.markdown('<div class="header-btn-container">', unsafe_allow_html=True)
        if st.button(
            "➕ เพิ่มผู้รับจ้าง", use_container_width=True, key="hiring_btn_add_shop"
        ):
            add_shop_modal()
        st.markdown("</div>", unsafe_allow_html=True)

    with top_col4:
        st.markdown('<div class="header-btn-container">', unsafe_allow_html=True)
        if st.button(
            "✏️ แก้ไขร้าน", use_container_width=True, key="hiring_btn_edit_shop"
        ):
            edit_hiring_shop_dialog(shop_list_options, df_shops)
        st.markdown("</div>", unsafe_allow_html=True)

    st.write("")
    col_ratios = [3, 2, 1, 5]
    h1, h2, h3, h4 = st.columns(col_ratios, gap="small")
    h1.caption("**ชื่อผู้รับจ้าง/ร้านค้า**")
    h2.caption("**วงเงิน**")
    h3.caption("**จำนวน**")
    h4.caption("**รายละเอียดผู้รับจ้าง**")

    selected_vendors, budgetmid_list, vendor_item_counts, vendor_details_list = (
        [],
        [],
        [],
        [],
    )

    def on_vendor_change(index):
        detail_key = f"hiring_vendor_detail_{index}"
        if detail_key in st.session_state:
            del st.session_state[detail_key]

    for i in range(shop_count):
        c1, c2, c3, c4 = st.columns(col_ratios, gap="small")

        with c1:
            vendor = st.selectbox(
                f"ผู้รับจ้าง {i+1}",
                options=shop_list_options,
                index=0 if shop_list_options else None,
                key=f"hiring_selected_vendor_name_{i+1}",
                on_change=on_vendor_change,
                args=(i + 1,),
                label_visibility="collapsed",
            )
            selected_vendors.append(vendor)

        with c2:
            b_mid = st.text_input(
                f"วงเงิน {i+1}",
                placeholder="ตัวอย่าง: 15000",
                key=f"hiring_budget_mid_{i+1}",
                label_visibility="collapsed",
            )
            budgetmid_list.append(b_mid)

        with c3:
            v_item_count = st.text_input(
                f"จำนวน {i+1}",
                value=item_count_raw if (item_count_raw and shop_count == 1) else "",
                placeholder="",
                key=f"hiring_item_count_{i+1}",
                label_visibility="collapsed",
            )
            vendor_item_counts.append(v_item_count)

        with c4:
            detail_lines = []
            c_addr, c_phone, c_tax = "", "", ""
            if vendor and not df_shops.empty:
                match = df_shops[df_shops["clean_shop_name"] == str(vendor).strip()]
                if not match.empty:
                    c_addr, c_phone, c_tax = _extract_shop_info(match.iloc[0])
                    if c_addr:
                        detail_lines.append(f"ที่อยู่: {c_addr}")
                    if c_phone:
                        detail_lines.append(f"โทร: {c_phone}")
                    if c_tax:
                        detail_lines.append(f"เลขภาษี: {c_tax}")
                    detail_text = (
                        "\n".join(detail_lines)
                        if detail_lines
                        else "ไม่มีข้อมูลเพิ่มเติม"
                    )
                else:
                    detail_text = "ไม่พบข้อมูลในระบบ"
            else:
                detail_text = ""

            st.text_area(
                f"รายละเอียด {i+1}",
                value=detail_text,
                disabled=True,
                height=80,
                label_visibility="collapsed",
            )
            vendor_details_list.append((c_addr, c_phone, c_tax))

    # --- ส่วนที่ 4: ใบสั่งจ้าง ---
    st.write("")
    st.write("")
    st.subheader("📄 4. ใบสั่งจ้าง / ใบข้อตกลงจ้าง")
    col1, col2 = st.columns([1, 1], gap="large")
    submit_no = col1.text_input(
        "เลขที่ข้อตกลง", placeholder="ตัวอย่าง: 961/2569", key="hiring_submit_no"
    )
    default_order_date = add_business_days(selected_date2, 3)
    selected_date5 = col2.date_input(
        "วันที่ใบสั่งจ้าง", value=default_order_date, key="hiring_date_order_input"
    )

    # --- ส่วนที่ 5: ใบตรวจรับการจ้าง ---
    st.write("")
    st.subheader("✅ 5. ใบตรวจรับการจ้าง")
    selected_date4 = st.date_input(
        "วันที่ตรวจรับ", datetime.date.today(), key="hiring_check_date"
    )

    # --- ส่วนที่ 6: รายละเอียดรายการพัสดุและสเปค (Data Editor เสถียร ไม่รีเซ็ตซ้ำซ้อน) ---
    st.write("")
    st.subheader("📋 6. รายละเอียดรายการพัสดุและสเปค (Item Specifications)")
    st.caption(
        "กำหนดรายการพัสดุหรือขอบเขตงานจ้าง ระบบจะนำไปสร้างไฟล์ Excel ข้อกำหนดการจ้าง (Space) ให้อัตโนมัติเมื่อกดสร้างเอกสาร"
    )

    edited_items_df = st.data_editor(
        st.session_state["hiring_items_editor"],  # 👈 เติมตัวแปรข้อมูลตรงนี้ครับ
        num_rows="dynamic",
        use_container_width=True,
        key="hiring_items_editor",
        column_config={
            "รายการพัสดุ / รายละเอียดสเปค": st.column_config.TextColumn(
                "รายการพัสดุ / รายละเอียดสเปค / ขอบเขตงาน", width="large"
            ),
            "จำนวน": st.column_config.NumberColumn("จำนวน", width="small", format="%d"),
            "หน่วย": st.column_config.TextColumn("หน่วย", width="small"),
        },
    )

    # --- ประมวลผลแปลงตัวเลข/วันที่ ---
    formatted_budget = format_budget_money(budget, use_thai=use_thai_num)
    formatted_date = format_thai_date(selected_date, use_thai=use_thai_num)
    formatted_date2 = format_thai_date(selected_date2, use_thai=use_thai_num)
    calc_date3 = add_business_days(selected_date5, 3)
    formatted_date3 = format_thai_date(calc_date3, use_thai=use_thai_num)
    formatted_date4 = format_thai_date(selected_date4, use_thai=use_thai_num)
    formatted_date5 = format_thai_date(selected_date5, use_thai=use_thai_num)

    formatted_doc_no = to_thai_num(doc_no_raw) if use_thai_num else str(doc_no_raw)
    formatted_doc_no_save = (
        to_thai_num(doc_no_save) if use_thai_num else str(doc_no_save)
    )

    def generate_submit_no_for_vendor(base_submit_no, index):
        if not base_submit_no:
            return ""
        if index == 0:
            return to_thai_num(base_submit_no) if use_thai_num else str(base_submit_no)
        new_sub_no = (
            f"{base_submit_no.split('/', 1)[0]}.{index}/{base_submit_no.split('/', 1)[1]}"
            if "/" in base_submit_no
            else f"{base_submit_no}.{index}"
        )
        return to_thai_num(new_sub_no) if use_thai_num else str(new_sub_no)

    sub_nos = [generate_submit_no_for_vendor(submit_no, i) for i in range(4)]
    budget_with_text = f"{formatted_budget} บาท ({budget_text})"

    def get_vendor_info(idx):
        v_name = selected_vendors[idx] if len(selected_vendors) > idx else ""
        v_b_mid = budgetmid_list[idx] if len(budgetmid_list) > idx else ""
        v_cnt = vendor_item_counts[idx] if len(vendor_item_counts) > idx else ""
        c_addr, c_phone, c_tax = (
            vendor_details_list[idx] if len(vendor_details_list) > idx else ("", "", "")
        )

        c_num_mid = 0.0
        try:
            c_num_mid = float(
                v_b_mid.replace(",", "").replace(" ", "").replace("บาท", "")
            )
            b_text_mid = bahttext(c_num_mid)
        except ValueError:
            b_text_mid = v_b_mid if v_b_mid else "-"

        c_cnt_num = 0
        try:
            match_item = re.search(r"\d+", str(v_cnt).replace(",", ""))
            if match_item:
                c_cnt_num = int(match_item.group())
        except ValueError:
            c_cnt_num = 0

        fmt_b_mid = (
            format_budget_money(v_b_mid, use_thai=use_thai_num2) if v_b_mid else ""
        )
        b_with_text_mid = f"{fmt_b_mid} บาท ({b_text_mid})" if fmt_b_mid else ""
        fmt_item_cnt = to_thai_num(v_cnt) if (use_thai_num and v_cnt) else str(v_cnt)

        return {
            "name": v_name,
            "addr": to_thai_num(c_addr) if use_thai_num else c_addr,
            "phone": to_thai_num(c_phone) if use_thai_num else c_phone,
            "tax": to_thai_num(c_tax) if use_thai_num else c_tax,
            "budget_mid": fmt_b_mid,
            "budget_text_mid": b_text_mid,
            "budget_with_text_mid": b_with_text_mid,
            "item_count": fmt_item_cnt,
            "clean_num_mid": c_num_mid,
            "clean_cnt_num": c_cnt_num,
        }

    v_list = [get_vendor_info(i) for i in range(4)]
    v1, v2, v3, v4 = v_list

    # --- คำนวณผลรวม ---
    total_budget_mid_num = sum(v["clean_num_mid"] for v in v_list[:shop_count])
    total_item_count_num = sum(v["clean_cnt_num"] for v in v_list[:shop_count])

    formatted_total_budget_mid = format_budget_money(
        str(total_budget_mid_num), use_thai=use_thai_num2
    )
    total_budget_text_mid = (
        bahttext(total_budget_mid_num) if total_budget_mid_num > 0 else "-"
    )
    total_budget_with_text_mid = (
        f"{formatted_total_budget_mid} บาท ({total_budget_text_mid})"
        if total_budget_mid_num > 0
        else ""
    )
    formatted_total_item_count = (
        to_thai_num(str(total_item_count_num))
        if use_thai_num
        else str(total_item_count_num)
    )
    formatted_project_name = to_thai_num(project_name) if use_thai_num else project_name

    # --- รวบรวมข้อมูล replacement สำหรับ Docx ---
    replacements_data = {
        "{{PROJECT_NAME}}": formatted_project_name,
        "{{BUDGET}}": formatted_budget,
        "{{BUDGET_TEXT}}": budget_text,
        "{{BUDGET_WITH_TEXT}}": budget_with_text,
        "{{BUDGET_TYPE}}": budget_type_text,
        "{{DEPARTMENT}}": department,
        "{{ITEM_COUNT}}": v1["item_count"],
        "{{BUDGET_MID}}": v1["budget_mid"],
        "{{BUDGET_TEXT_MID}}": v1["budget_text_mid"],
        "{{BUDGET_WITH_TEXT_MID}}": v1["budget_with_text_mid"],
        "{{TOTAL_ITEM_COUNT}}": formatted_total_item_count,
        "{{TOTAL_BUDGET_MID}}": formatted_total_budget_mid,
        "{{TOTAL_BUDGET_TEXT_MID}}": total_budget_text_mid,
        "{{TOTAL_BUDGET_WITH_TEXT_MID}}": total_budget_with_text_mid,
        "{{VENDOR_NAME}}": v1["name"],
        "{{VENDOR_ADDRESS}}": v1["addr"],
        "{{VENDOR_PHONE}}": v1["phone"],
        "{{VENDOR_TAX_ID}}": v1["tax"],
        "{{SUBMIT_NO}}": sub_nos[0],
        "{{PERCEL_NO}}": parcel_no,
        "{{DOC_DATE}}": formatted_date,
        "{{DOC_DATE2}}": formatted_date2,
        "{{DOC_DATE3}}": formatted_date3,
        "{{DOC_DATE4}}": formatted_date4,
        "{{DOC_DATE5}}": formatted_date5,
        "{{DOC_NO}}": formatted_doc_no,
        "{{DOC_NO_SAVE}}": formatted_doc_no_save,
        "{{DIRECTOR_INSPECTOR}}": "คณะกรรมการ" if check_count > 1 else "ผู้",
    }

    for idx, key_suffix in enumerate(["2", "3", "4"], start=1):
        v = v_list[idx]
        replacements_data.update(
            {
                f"{{{{VENDOR_NAME{key_suffix}}}}}": v["name"],
                f"{{{{VENDOR_ADDRESS{key_suffix}}}}}": v["addr"],
                f"{{{{VENDOR_PHONE{key_suffix}}}}}": v["phone"],
                f"{{{{VENDOR_TAX_ID{key_suffix}}}}}": v["tax"],
                f"{{{{BUDGET_MID{key_suffix}}}}}": v["budget_mid"],
                f"{{{{BUDGET_TEXT_MID{key_suffix}}}}}": v["budget_text_mid"],
                f"{{{{BUDGET_WITH_TEXT_MID{key_suffix}}}}}": v["budget_with_text_mid"],
                f"{{{{ITEM_COUNT{key_suffix}}}}}": v["item_count"],
                f"{{{{SUBMIT_NO{key_suffix}}}}}": sub_nos[idx],
            }
        )

    thai_nums = ["๑", "๒", "๓"]
    for idx, (name, acad, pos) in enumerate(buy_persons):
        suf = "" if idx == 0 else str(idx + 1)
        valid = name and (buy_count >= idx + 1)
        replacements_data.update(
            {
                f"{{{{DIRECTOR_NAME_BUY{suf}}}}}": (
                    f"{thai_nums[idx]}. {name}" if valid else ""
                ),
                f"{{{{DIRECTOR_ACAD_BUY{suf}}}}}": acad if valid else "",
                f"{{{{DIRECTOR_POS_BUY{suf}}}}}": pos if valid else "",
                f"{{{{DIRECTOR_NAME_BUY{suf}_PLAIN}}}}": name if valid else "",
                f"{{{{SIGN_LINE_BUY{idx+1}}}}}": (
                    f"ลงชื่อ....................................{pos if pos else 'กรรมการฯ'}"
                    if valid
                    else ""
                ),
                f"{{{{SIGN_NAME_BUY{idx+1}}}}}": f"({name})" if valid else "",
            }
        )

    for idx, (name, acad, pos) in enumerate(check_persons):
        suf = "" if idx == 0 else str(idx + 1)
        valid = name and (check_count >= idx + 1)
        actual_pos = (
            ("ผู้ตรวจรับการจ้าง" if check_count == 1 else pos) if idx == 0 else pos
        )

        sign_line = (
            "ลงชื่อ....................................ผู้ตรวจรับการจ้าง"
            if (idx == 0 and check_count == 1)
            else (
                "ลงชื่อ....................................ประธานกรรมการฯ"
                if idx == 0
                else "ลงชื่อ....................................กรรมการฯ"
            )
        )

        replacements_data.update(
            {
                f"{{{{CHECKITEM_NAME{suf}}}}}": (
                    f"{thai_nums[idx]}. {name}" if valid else ""
                ),
                f"{{{{CHECKITEM_ACAD{suf}}}}}": acad if valid else "",
                f"{{{{CHECKITEM_POS{suf}}}}}": actual_pos if valid else "",
                f"{{{{CHECKITEM_NAME{suf}_PLAIN}}}}": name if valid else "",
                f"{{{{SIGN_LINE_CHECK{idx+1}}}}}": sign_line if valid else "",
                f"{{{{SIGN_NAME_CHECK{idx+1}}}}}": f"({name})" if valid else "",
            }
        )

    # --- ส่วนที่ 7: ตรวจสอบข้อมูลสรุป (Preview), Validation & Actions ---
    st.write("")
    st.subheader("👁️ 7. ตรวจสอบข้อมูลสรุป (Preview)")

    with st.container(border=True):
        st.markdown("##### 📄 สรุปรายละเอียดเอกสารจัดจ้าง")
        col_prev1, col_prev2 = st.columns([1, 1], gap="medium")

        with col_prev1:
            st.markdown(
                f"**โครงการ:** {project_name if project_name else '⚠️ *(ยังไม่ได้กรอก)*'}"
            )
            st.markdown(
                f"**หน่วยงาน/แผนก:** {department if department else '⚠️ *(ยังไม่ได้กรอก)*'}"
            )
            st.markdown(
                f"**วงเงินงบประมาณ:** {budget_with_text if budget else '⚠️ *(ยังไม่ได้กรอก)*'}"
            )

            if clean_num > 0 and total_budget_mid_num > clean_num:
                st.markdown(
                    f"**ราคากลางรวม ({shop_count} ผู้รับจ้าง):** {formatted_total_budget_mid} บาท 🚨 **(สูงกว่าวงเงินงบประมาณ)**"
                )
            elif total_budget_mid_num > 0:
                st.markdown(
                    f"**ราคากลางรวม ({shop_count} ผู้รับจ้าง):** {formatted_total_budget_mid} บาท"
                )

            st.markdown(f"**แหล่งเงิน:** {budget_type_text}")

            if clean_max_item_count > 0 and total_item_count_num > clean_max_item_count:
                st.markdown(
                    f"**จำนวนรายการรวม ({shop_count} ผู้รับจ้าง):** {formatted_total_item_count} รายการ 🚨 **(เกินจำนวนรายการทั้งหมด {to_thai_num(str(clean_max_item_count))} รายการ)**"
                )
            else:
                st.markdown(
                    f"**จำนวนรายการทั้งหมด:** {to_thai_num(item_count_raw) if item_count_raw else '⚠️ *(ยังไม่ได้กรอก)*'}"
                )

        with col_prev2:
            st.markdown(
                f"**เลขพัสดุที่ใช้ระบุชื่อไฟล์:** `{parcel_no if parcel_no else 'ไม่ได้ระบุ (ใช้ตามชื่อต้นแบบ)'}`"
            )
            st.markdown(
                f"**เลขที่คำสั่ง / ข้อตกลง:** {formatted_doc_no if doc_no_raw else '-'} / {sub_nos[0] if submit_no else '-'}"
            )
            st.markdown(
                f"**ผู้รับจ้าง / ร้านค้า:** {v1['name'] if v1['name'] else '⚠️ *(ยังไม่ได้เลือก)*'}"
            )
            st.markdown(f"**วันที่เอกสาร (ส.1):** {formatted_date}")
            st.markdown(f"**วันที่รายงานผล:** {formatted_date2}")
            st.markdown(f"**วันที่ใบสั่งจ้าง:** {formatted_date5}")
            st.markdown(f"**วันที่ตรวจรับ:** {formatted_date4}")

        st.markdown("---")
        col_prev_buy, col_prev_check = st.columns([1, 1], gap="medium")

        with col_prev_buy:
            st.markdown("**🛠️ รายชื่อคณะกรรมการจัดจ้าง:**")
            buy_list_preview = [
                f"{i+1}. {p[0]} ({p[2]})"
                for i, p in enumerate(buy_persons[:buy_count])
                if p[0]
            ]
            if buy_list_preview:
                for item in buy_list_preview:
                    st.text(f"  • {item}")
            else:
                st.caption("⚠️ ยังไม่มีการเลือกกรรมการจัดจ้าง")

        with col_prev_check:
            st.markdown("**🔍 รายชื่อคณะกรรมการตรวจรับ:**")
            check_list_preview = [
                f"{i+1}. {p[0]} ({'ผู้ตรวจรับการจ้าง' if check_count == 1 and i == 0 else p[2]})"
                for i, p in enumerate(check_persons[:check_count])
                if p[0]
            ]
            if check_list_preview:
                for item in check_list_preview:
                    st.text(f"  • {item}")
            else:
                st.caption("⚠️ ยังไม่มีการเลือกกรรมการตรวจรับ")

    st.write("")
    col_action1, col_action2 = st.columns([1, 4], gap="small")

    with col_action1:
        st.button(
            "🗑️ ล้างข้อมูล",
            use_container_width=True,
            help="ล้างข้อมูลการกรอกทั้งหมด",
            key="hiring_reset_btn",
            on_click=reset_hiring_form,
        )

    with col_action2:
        btn_generate = st.button(
            "🚀 สร้างเอกสารจัดจ้างทั้งหมด",
            type="primary",
            use_container_width=True,
            key="hiring_submit_btn",
        )

    if btn_generate:
        missing_fields = []
        if not project_name.strip():
            missing_fields.append("ชื่องานจ้าง / โครงการ")
        if not budget.strip():
            missing_fields.append("จำนวนเงิน / วงเงิน")
        if not department.strip():
            missing_fields.append("งาน หรือ แผนกวิชา")
        if not v1["name"]:
            missing_fields.append("ชื่อบริษัท / ผู้รับจ้าง")
        if not buy_persons[0][0]:
            missing_fields.append("ประธาน/กรรมการจัดจ้างคนที่ 1")
        if not check_persons[0][0]:
            missing_fields.append("ผู้ตรวจรับ / ประธานกรรมการตรวจรับคนที่ 1")

        if missing_fields:
            st.error(
                "❌ กรุณากรอกข้อมูลสำคัญให้ครบถ้วนก่อนสร้างเอกสาร:\n- "
                + "\n- ".join(missing_fields)
            )
        elif clean_num > 0 and total_budget_mid_num > clean_num:
            st.error(
                f"❌ ไม่สามารถสร้างเอกสารได้เนื่องจาก **ผลรวมราคากลางของผู้รับจ้างสูงกว่าวงเงินงบประมาณ** ({formatted_total_budget_mid} บาท > {formatted_budget} บาท) กรุณาตรวจสอบและแก้ไขข้อมูล"
            )
        elif clean_max_item_count > 0 and total_item_count_num > clean_max_item_count:
            st.error(
                f"❌ ไม่สามารถสร้างเอกสารได้เนื่องจาก **ผลรวมจำนวนรายการของผู้รับจ้าง เกินกว่าจำนวนรายการทั้งหมดที่ระบุไว้** ({formatted_total_item_count} รายการ > {to_thai_num(str(clean_max_item_count))} รายการ) กรุณาตรวจสอบและแก้ไขข้อมูล"
            )
        elif not template_files:
            st.error(
                f"ไม่สามารถสร้างเอกสารได้ เนื่องจากไม่มีไฟล์ต้นแบบในโฟลเดอร์ '{TEMPLATE_DIR}'"
            )
        else:
            with st.spinner(
                "กำลังประมวลผลเอกสารจัดจ้างและไฟล์ Excel ข้อกำหนดการจ้าง..."
            ):
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, "w") as zip_file:
                    # 1. สร้างและแพ็คไฟล์ Word (.docx)
                    for file_name in template_files:
                        file_path = os.path.join(TEMPLATE_DIR, file_name)
                        processed_stream = process_docx(
                            file_path,
                            replacements_data,
                            shop_count=shop_count,
                            buy_count=buy_count,
                            check_count=check_count,
                        )
                        new_filename = (
                            re.sub(r"\d+-\d+", parcel_no.strip(), file_name)
                            if parcel_no.strip()
                            else file_name
                        )
                        zip_file.writestr(
                            f"{new_filename}", processed_stream.getvalue()
                        )

                    # 2. สร้างและแพ็คไฟล์ Excel ข้อกำหนดการจ้าง (Space) รวมลงใน ZIP เดียวกันทันที
                    valid_items_df = edited_items_df[
                        edited_items_df["รายการพัสดุ / รายละเอียดสเปค"]
                        .astype(str)
                        .str.strip()
                        != ""
                    ].copy()

                    excel_stream = generate_hiring_space_excel_internal(
                        valid_items=valid_items_df,
                        total_amount=clean_num if clean_num > 0 else 0.0,
                        project_name=project_name,
                        department=department,
                        budget_type=budget_type_text,
                        parcel_no=parcel_no,
                        receiver=buy_persons[0][0] if buy_persons else "",
                        receiver_sub=buy_persons[1][0] if len(buy_persons) > 1 else "",
                    )

                    clean_filename_doc_no = (
                        parcel_no.strip()
                        if parcel_no.strip()
                        else (
                            doc_no_raw.replace("/", "-") if doc_no_raw else "ไม่ระบุเลข"
                        )
                    )
                    excel_filename = (
                        f"เอกสารข้อกำหนดการจ้าง_Space_{clean_filename_doc_no}.xlsx"
                    )
                    zip_file.writestr(excel_filename, excel_stream.getvalue())

                zip_buffer.seek(0)

            st.success(
                "🎉 สร้างชุดเอกสารจัดจ้างและไฟล์ Excel ข้อกำหนดการจ้างเรียบร้อยแล้ว!"
            )
            st.download_button(
                label="📦 ดาวน์โหลดชุดเอกสารจัดจ้างและ Excel (.zip)",
                data=zip_buffer,
                file_name=f"เอกสารจัดจ้าง_{project_name}_{clean_filename_doc_no}.zip",
                mime="application/zip",
                use_container_width=True,
            )


if __name__ == "__main__":
    render_hiring_page()
