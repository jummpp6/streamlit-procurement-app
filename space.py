# ชื่อไฟล์: space.py
from copy import copy
import io
import os
from google.oauth2.service_account import Credentials
import gspread
import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
import pandas as pd
from pythainlp.util import bahttext
import streamlit as st

# 🔗 กำหนด Spreadsheet ID ของ Google Sheet
SPREADSHEET_ID = "1k_hSSdF50uYcRZVffPNh0NfpXvbJFjajlG26SRUpYMs"


def get_gsheet_worksheet(sheet_name="Teachers"):
    """เชื่อมต่อ Google Sheets ด้วย Service Account ผ่าน gspread"""
    scopes = ["https://www.googleapis.com/auth/spreadsheets"]
    creds = Credentials.from_service_account_info(
        st.secrets["gcp_service_account"], scopes=scopes
    )
    client = gspread.authorize(creds)
    sheet = client.open_by_key(SPREADSHEET_ID).worksheet(sheet_name)
    return sheet


@st.cache_data(ttl=60)
def get_only_teacher_names(file_path: str = "Teachers") -> list[str]:
    """ดึงรายชื่อผู้รับพัสดุ/ครูจาก Google Sheets"""
    default_names = ["นายสมชาย ใจดี", "นางสาววิภา รักสงบ"]
    try:
        sheet_name = "Teachers"
        if isinstance(file_path, str) and file_path.endswith(".xlsx"):
            sheet_name = "Teachers"
        elif isinstance(file_path, str) and file_path.strip():
            sheet_name = file_path

        ws = get_gsheet_worksheet(sheet_name)
        rows = ws.get_all_values()

        clean_names = []
        if len(rows) > 1:
            for r in rows[1:]:
                if len(r) >= 2:
                    full_name = str(r[1]).strip() if r[1] else ""
                    if (
                        full_name
                        and full_name
                        not in [
                            "ชื่อ - สกุล",
                            "ชื่อ-สกุล",
                            "ชื่อ",
                            "ลำดับที่",
                            "nan",
                            "None",
                        ]
                        and full_name not in clean_names
                    ):
                        clean_names.append(full_name)

        if clean_names:
            return clean_names
    except Exception as e:
        st.warning(f"⚠️ ไม่สามารถอ่านรายชื่อจาก Google Sheets ได้: {e}")

    return default_names


def generate_fourcolor_excel(
    template_path,
    receiver,
    receiver_sub,
    valid_items,
    total_amount,
    project_name="",
    **kwargs,
):
    wb = openpyxl.load_workbook(template_path)

    try:
        budget_text = bahttext(total_amount)
    except Exception:
        budget_text = ""

    budget_with_text = (
        f"{total_amount:,.2f} บาท ({budget_text})" if total_amount > 0 else ""
    )

    mapping = {
        "{{BUDGET_WITH_TEXT}}": budget_with_text,
        "{{budget_with_text}}": budget_with_text,
        "{{BUDGET_TEXT_MID}}": budget_text,
        "{{budget_text_mid}}": budget_text,
        "{{receiver}}": str(receiver),
        "{{receiver_sub}}": str(receiver_sub),
        "{{total_amount}}": f"{total_amount:,.2f}",
    }

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    empty_border = Border()

    for ws in wb.worksheets:
        start_row = None

        for r in range(1, ws.max_row + 1):
            val_col1 = str(ws.cell(row=r, column=1).value or "")
            val_col2 = str(ws.cell(row=r, column=2).value or "")
            if "{% for item in items %}" in val_col1 or "{{item.name}}" in val_col2:
                start_row = r
                break

        if start_row:
            num_items = len(valid_items)
            is_space_sheet = ws.title == "Space"

            template_styles = {}
            max_cols = 4 if is_space_sheet else 8
            for c in range(1, max_cols + 1):
                src_cell = ws.cell(row=start_row, column=c)
                template_styles[c] = {
                    "font": copy(src_cell.font),
                    "alignment": copy(src_cell.alignment),
                }

            merged_to_shift = []
            merged_to_remove = []
            for rng in list(ws.merged_cells.ranges):
                if rng.min_row > start_row:
                    merged_to_shift.append(
                        (rng.min_row, rng.min_col, rng.max_row, rng.max_col)
                    )
                    merged_to_remove.append(str(rng))

            for rng_str in merged_to_remove:
                ws.unmerge_cells(rng_str)

            num_page_breaks = (num_items - 1) // 19 if num_items > 19 else 0
            shift_amount = (num_items - 1) + (num_page_breaks * 2)

            if shift_amount > 0:
                ws.insert_rows(start_row + 1, amount=shift_amount)

            for min_r, min_c, max_r, max_c in merged_to_shift:
                ws.merge_cells(
                    start_row=min_r + shift_amount,
                    start_column=min_c,
                    end_row=max_r + shift_amount,
                    end_column=max_c,
                )

            current_row = start_row
            for idx, item in valid_items.reset_index(drop=True).iterrows():
                if is_space_sheet:
                    ws.cell(row=current_row, column=1, value=idx + 1)
                    ws.cell(row=current_row, column=2, value=item["name"])
                    ws.cell(row=current_row, column=3, value=item["quantity"])
                    ws.cell(row=current_row, column=4, value=item["unit"])

                    for col in range(1, 5):
                        cell = ws.cell(row=current_row, column=col)
                        cell.border = thin_border
                        if template_styles.get(col):
                            cell.font = copy(template_styles[col]["font"])
                            cell.alignment = copy(template_styles[col]["alignment"])

                    ws.cell(row=current_row, column=1).alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                    ws.cell(row=current_row, column=3).alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                    ws.cell(row=current_row, column=4).alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                else:
                    ws.merge_cells(
                        start_row=current_row,
                        start_column=2,
                        end_row=current_row,
                        end_column=4,
                    )
                    ws.cell(row=current_row, column=1, value=idx + 1)
                    ws.cell(row=current_row, column=2, value=item["name"])
                    ws.cell(row=current_row, column=5, value=item["quantity"])
                    ws.cell(row=current_row, column=7, value=item["unit"])

                    for col in range(1, 9):
                        cell = ws.cell(row=current_row, column=col)
                        cell.border = thin_border
                        if template_styles.get(col):
                            cell.font = copy(template_styles[col]["font"])

                    ws.cell(row=current_row, column=1).alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                    ws.cell(row=current_row, column=5).alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                    ws.cell(row=current_row, column=7).alignment = Alignment(
                        horizontal="center", vertical="center"
                    )

                if (idx + 1) % 19 == 0 and (idx + 1) < num_items:
                    max_c = 4 if is_space_sheet else 8

                    current_row += 1
                    for col in range(1, max_c + 1):
                        ws.cell(row=current_row, column=col).border = empty_border

                    current_row += 1
                    next_page_num = ((idx + 1) // 19) + 1

                    ws.merge_cells(
                        start_row=current_row,
                        start_column=1,
                        end_row=current_row,
                        end_column=max_c,
                    )
                    cell_continue = ws.cell(row=current_row, column=1)
                    cell_continue.value = f"(ต่อแผ่น {next_page_num})"
                    cell_continue.alignment = Alignment(
                        horizontal="right", vertical="center"
                    )

                    for col in range(1, max_c + 1):
                        ws.cell(row=current_row, column=col).border = empty_border

                    if template_styles.get(1):
                        base_font = copy(template_styles[1]["font"])
                        base_font.italic = True
                        base_font.bold = True
                        cell_continue.font = base_font

                    ws.row_breaks.append(Break(id=current_row))

                current_row += 1

        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell_obj = ws.cell(row=r, column=c)
                if type(cell_obj).__name__ != "MergedCell":
                    val = str(cell_obj.value or "")

                    if "{{PROJECT_NAME}}" in val or "{{project_name}}" in val:
                        cell_obj.value = val.replace(
                            "{{PROJECT_NAME}}", str(project_name)
                        ).replace("{{project_name}}", str(project_name))

                        current_align = cell_obj.alignment
                        cell_obj.alignment = Alignment(
                            horizontal=current_align.horizontal or "left",
                            vertical=current_align.vertical or "center",
                            wrap_text=True,
                        )

                        name_len = len(str(project_name))
                        if name_len > 90:
                            ws.row_dimensions[r].height = 55
                        elif name_len > 45:
                            ws.row_dimensions[r].height = 42
                    else:
                        for key, target in mapping.items():
                            if key in val:
                                cell_obj.value = val.replace(key, target)

        max_col_letter = get_column_letter(ws.max_column)
        ws.print_area = f"A1:{max_col_letter}{ws.max_row}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_space_excel(*args, **kwargs):
    """Alias สำหรับรองรับการเรียกใช้งานจาก space_fourcolor.py"""
    return generate_fourcolor_excel(*args, **kwargs)


@st.dialog("🎨 กรอกข้อมูลเอกสารสี่สี / คุณลักษณะสินค้า", width="large")
def render_fourcolor_dialog(
    default_receiver=None,
    default_receiver_sub=None,
    default_items=None,
    default_total_amount=0.0,
    project_name="",
    **kwargs,
):
    st.markdown(
        """
        <style>
            div[data-testid="stDialog"] div[data-baseweb="modal"] {
                max-width: 900px !important;
            }
        </style>
        """,
        unsafe_allow_html=True,
    )

    receiver_options = get_only_teacher_names("Teachers")

    st.markdown("##### 👤 ข้อมูลผู้รับพัสดุ")

    idx_rec = (
        receiver_options.index(default_receiver)
        if default_receiver in receiver_options
        else 0
    )
    idx_sub = (
        receiver_options.index(default_receiver_sub)
        if default_receiver_sub in receiver_options
        else (1 if len(receiver_options) > 1 else 0)
    )

    col_rec1, col_rec2 = st.columns(2)
    with col_rec1:
        receiver = st.selectbox(
            "ผู้รับพัสดุ (หลัก)",
            options=receiver_options,
            index=idx_rec,
            key="fourcolor_receiver",
        )
    with col_rec2:
        receiver_sub = st.selectbox(
            "ผู้รับพัสดุแทน",
            options=receiver_options,
            index=idx_sub,
            key="fourcolor_receiver_sub",
        )

    st.markdown("---")
    st.write(
        "กรอกรายการสินค้า จำนวน และราคา เพื่อใช้สำหรับสร้างเอกสารแนบสี่สี/คุณลักษณะ"
    )

    if "fourcolor_items_table" not in st.session_state:
        if default_items is not None and not default_items.empty:
            st.session_state["fourcolor_items_table"] = default_items.copy()
        else:
            st.session_state["fourcolor_items_table"] = pd.DataFrame(
                [
                    {
                        "name": "",
                        "quantity": 1.0,
                        "unit": "รายการ",
                        "price_per_unit": float(default_total_amount),
                    }
                ]
            )

    edited_df = st.data_editor(
        st.session_state["fourcolor_items_table"],
        num_rows="dynamic",
        use_container_width=True,
        hide_index=True,
        key="fourcolor_items_table",
        column_config={
            "name": st.column_config.Column(
                "รายการ",
                width="large",
                required=True,
            ),
            "quantity": st.column_config.NumberColumn(
                "จำนวน",
                min_value=0.0,
                step=1.0,
                default=1.0,
                required=True,
            ),
            "unit": st.column_config.TextColumn(
                "หน่วยนับ",
                default="รายการ",
                required=True,
            ),
            "price_per_unit": st.column_config.NumberColumn(
                "ราคาต่อหน่วย (บาท)",
                min_value=0.0,
                step=10.0,
                format="%.2f",
                default=0.0,
                required=True,
            ),
        },
    )

    processed_df = edited_df.copy()
    processed_df["total_price"] = processed_df["quantity"].fillna(0) * processed_df[
        "price_per_unit"
    ].fillna(0)

    valid_items = processed_df[processed_df["name"].str.strip() != ""].copy()
    valid_items.reset_index(drop=True, inplace=True)

    total_amount = valid_items["total_price"].sum() if not valid_items.empty else 0.0

    st.markdown(f"##### 📊 สรุปรายการทั้งหมด ({len(valid_items)} รายการ)")

    preview_df = valid_items.copy()
    preview_df.insert(0, "ลำดับ", range(1, len(preview_df) + 1))
    preview_df = preview_df.rename(
        columns={
            "name": "รายการ",
            "quantity": "จำนวน",
            "unit": "หน่วยนับ",
            "price_per_unit": "ราคาต่อหน่วย",
            "total_price": "ราคาสินค้า",
        }
    )

    st.dataframe(
        preview_df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "ราคาต่อหน่วย": st.column_config.NumberColumn(format="%.2f บาท"),
            "ราคาสินค้า": st.column_config.NumberColumn(format="%.2f บาท"),
        },
    )

    st.info(f"💰 **ราคารวมทั้งหมด:** {total_amount:,.2f} บาท")

    is_price_matched = abs(total_amount - default_total_amount) < 0.01

    if not is_price_matched and default_total_amount > 0:
        st.warning(
            f"⚠️ ราคารวมในตาราง ({total_amount:,.2f} บาท) **ยังไม่ตรงกับ** ราคารวมของร้านค้า ({default_total_amount:,.2f} บาท)"
        )

    col_btn1, col_btn2 = st.columns([1, 1])
    with col_btn1:
        if st.button("❌ ยกเลิก / ปิด", use_container_width=True):
            if "fourcolor_items_table" in st.session_state:
                del st.session_state["fourcolor_items_table"]
            st.rerun()

    with col_btn2:
        template_file = (
            "fourcolor_template.xlsx"
            if os.path.exists("fourcolor_template.xlsx")
            else os.path.join("templates_4color", "fourcolor_template.xlsx")
        )

        if os.path.exists(template_file):
            excel_bytes = generate_fourcolor_excel(
                template_file,
                receiver,
                receiver_sub,
                valid_items,
                total_amount,
                project_name=project_name,
                **kwargs,
            )

            if default_total_amount > 0 and not is_price_matched:
                st.button(
                    "📥 ดาวน์โหลดเอกสาร (Excel)",
                    disabled=True,
                    type="primary",
                    use_container_width=True,
                    help="ราคารวมในตารางต้องเท่ากับราคารวมของร้านค้าจึงจะดาวน์โหลดได้",
                )
            else:
                st.download_button(
                    label="📥 ดาวน์โหลดเอกสาร (Excel)",
                    data=excel_bytes,
                    file_name="เอกสารคุณลักษณะสินค้า_สี่สี.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True,
                )
        else:
            st.error(f"⚠️ ไม่พบไฟล์ Template: `{template_file}`")
