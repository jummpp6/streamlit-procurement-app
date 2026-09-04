from copy import copy
import io
import re
import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
import pandas as pd


# ฟังก์ชันแปลงตัวเลขเป็นตัวอักษรภาษาไทย (Baht Text)
def bahttext(number):
    if number is None:
        return ""
    try:
        num = float(number)
    except Exception:
        return ""
    if num == 0:
        return "ศูนย์บาทถ้วน"
    txtnum = ["ศูนย์", "หนึ่ง", "สอง", "สาม", "สี่", "ห้า", "หก", "เจ็ด", "แปด", "เก้า"]
    txtpos = ["", "สิบ", "ร้อย", "พัน", "หมื่น", "แสน", "ล้าน"]

    def convert(n):
        if n == 0:
            return ""
        s = ""
        s_num = str(n)
        l = len(s_num)
        for i, c in enumerate(s_num):
            d = int(c)
            p = l - i - 1
            if d == 0:
                continue
            if p == 1 and d == 1:
                s += "สิบ"
            elif p == 1 and d == 2:
                s += "ยี่สิบ"
            elif p == 0 and d == 1 and l > 1 and int(s_num[-2]) > 0:
                s += "เอ็ด"
            else:
                s += txtnum[d] + txtpos[p]
        return s

    s_val = f"{num:.2f}"
    baht_str, satang_str = s_val.split(".")
    baht_val = int(baht_str)
    satang_val = int(satang_str)
    baht_txt = ""
    if baht_val == 0:
        baht_txt = "ศูนย์"
    else:
        millions = baht_val // 1000000
        remainder = baht_val % 1000000
        if millions > 0:
            baht_txt += convert(millions) + "ล้าน"
        if remainder > 0:
            baht_txt += convert(remainder)
    res = baht_txt + "บาท"
    if satang_val == 0:
        res += "ถ้วน"
    else:
        res += convert(satang_val) + "สตางค์"
    return res


def get_real_cell(ws, row, col):
    cell = ws.cell(row=row, column=col)
    if type(cell).__name__ == "MergedCell":
        coord = cell.coordinate
        for cr in ws.merged_cells.ranges:
            if coord in cr:
                return ws.cell(row=cr.min_row, column=cr.min_col)
    return cell


def set_cell_value_safe(ws, row, col, value):
    cell = get_real_cell(ws, row, col)
    cell.value = value
    return cell


def format_baht_satang(val):
    try:
        if pd.isna(val) or val == "":
            return "", "-"
        num = float(str(val).replace(",", ""))
        if num.is_integer():
            return int(num), "-"
        else:
            baht = int(num)
            satang = round((num - baht) * 100)
            return baht, f"{satang:02d}"
    except Exception:
        return val, "-"


def generate_fourcolor_excel(
    template_path,
    receiver,
    receiver_sub,
    valid_items,
    total_amount,
    project_name="",
    department="",
    budget_type="",
    percel_no="",
):
    wb = openpyxl.load_workbook(template_path)
    budget_text = bahttext(total_amount)
    budget_with_text = (
        f"{total_amount:,.2f} บาท ({budget_text})" if total_amount > 0 else ""
    )

    formatted_parcel_no = str(percel_no)
    if formatted_parcel_no:
        formatted_parcel_no = re.sub(r"[.\-=_]", "/", formatted_parcel_no)
    else:
        formatted_parcel_no = "/"

    raw_mapping = {
        "BUDGET_WITH_TEXT": budget_with_text,
        "BUDGET_TEXT_MID": budget_text,
        "RECEIVER": str(receiver),
        "RECEIVER_SUB": str(receiver_sub),
        "RECEIVER_NO": str(receiver),
        "PROJECT_NAME": str(project_name),
        "DEPARTMENT": str(department),
        "BUDGET_TYPE": str(budget_type),
        "PERCEL_NO": formatted_parcel_no,
        "GRAND_TOTAL": f"{total_amount:,.2f}",
        "total_amount": f"{total_amount:,.2f}",
    }

    target_sheet_name = "Fourcolor"
    if target_sheet_name in wb.sheetnames:
        ws_template = wb[target_sheet_name]
    else:
        ws_template = wb.active

    if ws_template:
        limit_per_page = 19
        max_cols = 11
        num_items = len(valid_items)
        total_pages = max(1, (num_items + limit_per_page - 1) // limit_per_page)
        start_row = 11
        end_table_row = 29

        for page_idx in range(total_pages):
            if total_pages == 1:
                ws = ws_template
                ws.title = target_sheet_name
            else:
                if page_idx == 0:
                    ws = ws_template
                    ws.title = f"{target_sheet_name}_1"
                else:
                    ws = wb.copy_worksheet(ws_template)
                    ws.title = f"{target_sheet_name}_{page_idx + 1}"

            start_idx = page_idx * limit_per_page
            end_idx = min(start_idx + limit_per_page, num_items)
            page_items = valid_items.iloc[start_idx:end_idx]

            current_row = start_row
            for local_i, (_, item) in enumerate(
                page_items.reset_index(drop=True).iterrows()
            ):
                global_idx = start_idx + local_i
                item_name = str(item.get("name", ""))
                item_qty = float(item.get("quantity", 1.0))
                item_unit = str(item.get("unit", "รายการ"))
                item_price = float(item.get("price_per_unit", 0.0))
                item_total = item_qty * item_price

                ws.row_dimensions[current_row].height = 21
                row_template_styles = {}
                for c in range(1, max_cols + 1):
                    src_cell = ws_template.cell(row=start_row + local_i, column=c)
                    row_template_styles[c] = {
                        "font": copy(src_cell.font) if src_cell.font else None,
                        "alignment": (
                            copy(src_cell.alignment) if src_cell.alignment else None
                        ),
                        "border": copy(src_cell.border) if src_cell.border else None,
                    }

                u_baht, u_sat = format_baht_satang(item_price)
                t_baht, t_sat = format_baht_satang(item_total)

                set_cell_value_safe(ws, current_row, 1, global_idx + 1)
                set_cell_value_safe(ws, current_row, 3, item_name)
                set_cell_value_safe(ws, current_row, 4, "ป")
                set_cell_value_safe(ws, current_row, 5, item_unit)
                set_cell_value_safe(ws, current_row, 6, item_qty)
                set_cell_value_safe(ws, current_row, 8, u_baht)
                set_cell_value_safe(ws, current_row, 9, u_sat)
                set_cell_value_safe(ws, current_row, 10, t_baht)
                set_cell_value_safe(ws, current_row, 11, t_sat)

                for col in range(1, max_cols + 1):
                    cell = get_real_cell(ws, current_row, col)
                    style_data = row_template_styles.get(col)
                    if style_data:
                        if style_data.get("font"):
                            cell.font = copy(style_data["font"])
                        if style_data.get("border"):
                            cell.border = copy(style_data["border"])

                cell_col7 = get_real_cell(ws, current_row, 7)
                base_border = row_template_styles.get(7, {}).get("border")
                if base_border:
                    new_border = copy(base_border)
                    new_border.diagonal = Side(style="thin")
                    new_border.diagonalUp = True
                    cell_col7.border = new_border
                else:
                    cell_col7.border = Border(
                        left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="hair"),
                        bottom=Side(style="hair"),
                        diagonal=Side(style="thin"),
                        diagonalUp=True,
                    )

                get_real_cell(ws, current_row, 1).alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                get_real_cell(ws, current_row, 3).alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
                get_real_cell(ws, current_row, 4).alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                get_real_cell(ws, current_row, 5).alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                get_real_cell(ws, current_row, 6).alignment = Alignment(
                    horizontal="center", vertical="center"
                )
                current_row += 1

            summary_row = end_table_row + 1
            grand_summary_row = end_table_row + 2

            if page_idx == total_pages - 1 and current_row <= end_table_row:
                for col in range(1, max_cols + 1):
                    cell = get_real_cell(ws, current_row, col)
                    src_cell = ws_template.cell(row=current_row, column=col)
                    if src_cell.border:
                        new_border = copy(src_cell.border)
                        new_border.diagonal = None
                        new_border.diagonalUp = False
                        new_border.diagonalDown = False
                        cell.border = new_border
                    if col == 3:
                        cell_end = set_cell_value_safe(
                            ws, current_row, col, "หมดรายการ"
                        )
                        cell_end.font = Font(name="Angsana New", size=14, bold=True)
                        cell_end.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
                    else:
                        if col != 7:
                            set_cell_value_safe(ws, current_row, col, "")
                ws.row_dimensions[current_row].height = 21
                current_row += 1

            while current_row <= end_table_row:
                ws.row_dimensions[current_row].height = 21
                for col in range(1, max_cols + 1):
                    cell = get_real_cell(ws, current_row, col)
                    src_cell = ws_template.cell(row=current_row, column=col)
                    if src_cell.border:
                        new_border = copy(src_cell.border)
                        new_border.diagonal = None
                        new_border.diagonalUp = False
                        new_border.diagonalDown = False
                        cell.border = new_border
                    if col != 7:
                        set_cell_value_safe(ws, current_row, col, "")
                current_row += 1

            page_end_row = 10 + len(page_items)
            set_cell_value_safe(ws, summary_row, 4, "แผ่นนี้")
            set_cell_value_safe(ws, summary_row, 10, f"=SUM(J11:J{page_end_row})")
            set_cell_value_safe(ws, summary_row, 11, "-")

            page_cumulative_amount = (
                page_items["total_price"].sum()
                if total_pages == 1
                else valid_items.iloc[0:end_idx]["total_price"].sum()
            )
            page_budget_text = bahttext(page_cumulative_amount)

            set_cell_value_safe(ws, grand_summary_row, 4, "รวมทั้งสิ้น")
            set_cell_value_safe(ws, grand_summary_row, 6, page_budget_text)

            sum_col_letter = f"J{summary_row}"
            if total_pages == 1:
                set_cell_value_safe(ws, grand_summary_row, 10, f"={sum_col_letter}")
            else:
                if page_idx == 0:
                    set_cell_value_safe(ws, grand_summary_row, 10, f"={sum_col_letter}")
                else:
                    prev_sheet_name = f"{target_sheet_name}_{page_idx}"
                    set_cell_value_safe(
                        ws,
                        grand_summary_row,
                        10,
                        f"='{prev_sheet_name}'!J{grand_summary_row} + {sum_col_letter}",
                    )
            set_cell_value_safe(ws, grand_summary_row, 11, "-")

            for r_idx in range(11, 30):
                ws.row_dimensions[r_idx].height = 21

            for r in range(1, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    cell_obj = ws.cell(row=r, column=c)
                    if type(cell_obj).__name__ != "MergedCell" and cell_obj.value:
                        val = str(cell_obj.value)
                        changed = False
                        if "แผ่นที่" in val and "ของจำนวน" in val:
                            val = f"แผ่นที่                {page_idx + 1}          ของจำนวน             {total_pages}           แผ่น"
                            changed = True
                        elif "PROJECT_NAME" in val.upper():
                            val = re.sub(
                                r"\{\{\s*PROJECT_NAME\s*\}\}",
                                str(project_name),
                                val,
                                flags=re.IGNORECASE,
                            )
                            # จัดให้อยู่ตรงกลาง (ทั้งแนวนอนและแนวตั้ง) พร้อมเปิดย่อหน้าอัตโนมัติ
                            cell_obj.alignment = Alignment(
                                horizontal="center",
                                vertical="center",
                                wrap_text=True,
                                shrink_to_fit=True,
                            )

                            # ตรวจสอบความยาวเพื่อปรับขนาดฟอนต์และความสูงแถวอัตโนมัติหากชื่อโครงการยาว
                            p_len = len(str(project_name))
                            cur_font = cell_obj.font
                            f_name = (
                                cur_font.name
                                if cur_font and cur_font.name
                                else "Angsana New"
                            )
                            f_bold = cur_font.bold if cur_font else False

                            if p_len > 80:
                                cell_obj.font = Font(name=f_name, size=10, bold=f_bold)
                                ws.row_dimensions[r].height = 40
                            elif p_len > 50:
                                cell_obj.font = Font(name=f_name, size=12, bold=f_bold)
                                ws.row_dimensions[r].height = 32
                            elif p_len > 30:
                                cell_obj.font = Font(name=f_name, size=14, bold=f_bold)
                                ws.row_dimensions[r].height = 26

                            changed = True
                        else:
                            for key, target in raw_mapping.items():
                                pattern = r"\{\{\s*" + key + r"\s*\}\}"
                                if re.search(pattern, val, re.IGNORECASE):
                                    val = re.sub(
                                        pattern, str(target), val, flags=re.IGNORECASE
                                    )
                                    changed = True
                        if changed:
                            cell_obj.value = val

        max_col_letter = get_column_letter(max_cols)
        ws_template.print_area = f"A1:{max_col_letter}{ws_template.max_row}"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
